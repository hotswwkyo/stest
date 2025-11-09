# -*- coding: utf-8 -*-
import re
import json
import typing
import openpyxl
from openpyxl.cell import Cell


class ExcelDataFinder(object):
    """excel文件数据快速查找"""

    def __init__(self, filename, *, sheet_name=0, table_head_index=0, table_body_start_index=None, read_only=True, auto_open=True, **kwargs):

        self.filename = filename
        self.kwargs = kwargs
        self.read_only = read_only
        self.sheet_name = sheet_name
        self.table_head_index = table_head_index

        if not isinstance(table_body_start_index, int):
            self.table_body_start_index = self.table_head_index + 1
        else:
            self.table_body_start_index = table_body_start_index

        self.wb = None
        self.sheet = None
        self.end_row_index = None
        if auto_open:
            self.open()

    def __get_sheet(self, wb: openpyxl.Workbook, name_or_index):

        if isinstance(name_or_index, str):
            return wb[name_or_index]
        else:
            return wb.worksheets[name_or_index]

    def open(self):

        self.wb = openpyxl.load_workbook(self.filename, read_only=self.read_only, **self.kwargs)
        self.sheet = self.__get_sheet(self.wb, self.sheet_name)
        self.end_row_index = self.sheet.max_row

    def switch2sheet(self, name_or_index):

        self.sheet = self.__get_sheet(self.wb, name_or_index)
        self.sheet_name = name_or_index
        self.end_row_index = self.sheet.max_row

    def auto_set_index_by_table_head_title(self, *titles):

        hi = -1
        for index, row in enumerate(self.sheet.rows):
            titlelist = [cell.value for cell in row]
            has_not_exists = False
            for title in titles:
                if title not in titlelist:
                    has_not_exists = True
                    break
            if has_not_exists:
                continue
            hi = index
        if hi >= 0:
            self.table_head_index = hi
            self.table_body_start_index = self.table_head_index + 1
        return hi

    def __is_empty_row(self, row):

        blankcells = []
        for cell in row:
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                blankcells.append(True)
            else:
                blankcells.append(False)
        return all(blankcells)

    def set_end_row_index_by_empty_row(self):

        shengyu = list(self.sheet.rows)[self.table_head_index:]
        for i, row in enumerate(shengyu):
            if self.__is_empty_row(row):
                self.end_row_index = self.table_head_index + i
                break

    @property
    def other_table_head_rows(self):

        rows = []
        if self.table_head_index >= 1:
            rows = list(self.sheet.rows)[0:self.table_head_index]
        return rows

    @property
    def head(self):

        return list(self.sheet.rows)[self.table_head_index]

    @property
    def body(self):

        index = self.table_body_start_index
        return list(self.sheet.rows)[index:self.end_row_index]

    def search(self, row: dict, key_type: typing.Literal["column title", "column number", "column letter"] = "column title") -> typing.Optional[typing.Tuple[Cell]]:
        """返回匹配的第一行

        **Usage**

        Consider the following table:

        | ◥ | A | B | C | D | E |
        |:----:|:----|:-----|:-----|:-----|:-------|
        | **1** | 专资编码 | 院线 | 省 | 市 | 乙方签约协议 |
        | **2** | 80002212 | 辽宁北方 | 广东省 | 东莞市 | 已上传 |
        | **3** | 80002213 | 辽宁北方 | 北京市 | 北京市 | 未上传 |
        | **4** | 80002214 | 辽宁北方 | 新疆维吾尔自治区 | 乌鲁木齐市 | 已上传 |
        | **5** | 80002216 | 中影院线 | 宁夏回族自治区 | 银川市 | 已上传 |
        | **6** | 80002217 | 中影院线 | 广东省 | 东莞市 | 未上传 |

        ```python

        # Matches row 1
        checker.search({"省":"广东省","市":"东莞市"})
        checker.search({3:"广东省",4:"东莞市"}, key_type="column number")
        checker.search({"C":"广东省","D":"东莞市"}, key_type="column letter")

        # Matches row 4
        checker.search({"省":re.compile(\"自治区$\"),"乙方签约协议":"已上传"})
        checker.search({3:re.compile(\"自治区$\"),5:"已上传"}, key_type="column number")
        checker.search({"C":re.compile(\"自治区$\"),"E":"已上传"}, key_type="column letter")

        # Matches row 4
        def cell_matcher(k, cell):
            return True if re.compile("自治区$").search(cell.value) else False
        checker.search({"省":cell_matcher,"乙方签约协议":"已上传"})
        checker.search({3:cell_matcher,5:"已上传"}, key_type="column number")
        checker.search({"C":cell_matcher,"E":"已上传"}, key_type="column letter")

        ```

        Parameters
        ----------
        row : 行查询条件
            - key can be `column title` , `column number of cell (1-based)` or `column letter`
            - value 单元格内容、 re.Pattern对象，或 匹配函数，如果是匹配函数，则需要接收两个位置参数，第一个是行查询条件字典参数row的键，第二个是单元格对象(即openpyxl.cell.Cell实例)，如果匹配返回True，否则返回False

        key_type : 用于标记行查询条件字典参数row的键传的是 `column title`(列标题) `column number`(单元格列号) 还是 `column letter`(单元格列字母)，如果是`column title`，表格有多了标题是一样的，则按第一个匹配。默认是`column title`。如果表格有多个相同标题的列，建议使用`column number`(单元格列号) 或 `column letter`(单元格列字母)，因为这两个是唯一的。

        """

        right_row = None
        if len(self.body) < 1:
            return right_row
        if key_type == "column title":
            groups = self.group_header_cells_by_column_title()
        elif key_type == "column number":
            groups = self.group_header_cells_by_column_number()
        elif key_type == "column letter":
            groups = self.group_header_cells_by_column_letter()
        else:
            raise ValueError(f"Invalid parameter key_date value:f{key_type}")
        for brow in self.body:
            mismatch = False
            for k, v in row.items():
                cells = groups.get(k, None)
                if cells is None or len(cells) < 0:
                    mismatch = True
                else:
                    header_cell = cells[0]
                    cell_index = header_cell.column - 1
                    bcell = brow[cell_index]
                    if isinstance(v, re.Pattern):
                        if v.search(bcell.value) is None:
                            mismatch = True
                    elif callable(v):
                        if not v(k, bcell):
                            mismatch = True
                    else:
                        if bcell.value != v:
                            mismatch = True
                if mismatch:
                    break
            if not mismatch:
                right_row = brow
                break
        return right_row

    def findall(self, row: dict, key_type: typing.Literal["column title", "column number", "column letter"] = "column title") -> typing.List[typing.Tuple[Cell]]:
        """返回匹配的所有行

        **Usage**

        Consider the following table:

        | ◥ | A | B | C | D | E |
        |:----:|:----|:-----|:-----|:-----|:-------|
        | **1** | 专资编码 | 院线 | 省 | 市 | 乙方签约协议 |
        | **2** | 80002212 | 辽宁北方 | 广东省 | 东莞市 | 已上传 |
        | **3** | 80002213 | 辽宁北方 | 北京市 | 北京市 | 未上传 |
        | **4** | 80002214 | 辽宁北方 | 新疆维吾尔自治区 | 乌鲁木齐市 | 已上传 |
        | **5** | 80002216 | 中影院线 | 宁夏回族自治区 | 银川市 | 已上传 |
        | **6** | 80002217 | 中影院线 | 广东省 | 东莞市 | 未上传 |

        ```python

        # Matches row 1 and row 6
        checker.findall({"省":"广东省","市":"东莞市"})
        checker.findall({3:"广东省",4:"东莞市"}, key_type="column number")
        checker.findall({"C":"广东省","D":"东莞市"}, key_type="column letter")

        # Matches row 4 and row 5
        checker.findall({"省":re.compile(\"自治区$\"),"乙方签约协议":"已上传"})
        checker.findall({3:re.compile(\"自治区$\"),5:"已上传"}, key_type="column number")
        checker.findall({"C":re.compile(\"自治区$\"),"E":"已上传"}, key_type="column letter")

        # Matches row 4 and row 5
        def cell_matcher(k, cell):
            return True if re.compile("自治区$").search(cell.value) else False
        checker.findall({"省":cell_matcher,"乙方签约协议":"已上传"})
        checker.findall({3:cell_matcher,5:"已上传"}, key_type="column number")
        checker.findall({"C":cell_matcher,"E":"已上传"}, key_type="column letter")

        ```

        Parameters
        ----------
        row : 行查询条件
            - key can be `column title` , `column number of cell (1-based)` or `column letter`
            - value 单元格内容、 re.Pattern对象，或 匹配函数，如果是匹配函数，则需要接收两个位置参数，第一个是行查询条件字典参数row的键，第二个是单元格对象(即openpyxl.cell.Cell实例)，如果匹配返回True，否则返回False

        key_type : 用于标记行查询条件字典参数row的键传的是 `column title`(列标题) `column number`(单元格列号) 还是 `column letter`(单元格列字母)，如果是`column title`，表格有多了标题是一样的，则按第一个匹配。默认是`column title`，如果表格有多个相同标题的列，建议使用`column number`(单元格列号) 或 `column letter`(单元格列字母)，因为这两个是唯一的。

        """

        rowlist = []
        if len(self.body) < 1:
            return rowlist
        if key_type == "column title":
            groups = self.group_header_cells_by_column_title()
        elif key_type == "column number":
            groups = self.group_header_cells_by_column_number()
        elif key_type == "column letter":
            groups = self.group_header_cells_by_column_letter()
        else:
            raise ValueError(f"Invalid parameter key_date value:f{key_type}")
        for brow in self.body:
            mismatch = False
            for k, v in row.items():
                cells = groups.get(k, None)
                if cells is None or len(cells) < 0:
                    mismatch = True
                else:
                    header_cell = cells[0]
                    cell_index = header_cell.column - 1
                    bcell = brow[cell_index]
                    if isinstance(v, re.Pattern):
                        if v.search(bcell.value) is None:
                            mismatch = True
                    elif callable(v):
                        if not v(k, bcell):
                            mismatch = True
                    else:
                        if bcell.value != v:
                            mismatch = True
                if mismatch:
                    break
            if not mismatch:
                rowlist.append(brow)
        return rowlist

    def print_table_row(self, row: typing.Tuple[Cell]):

        for cell in row:
            print(cell.row, cell.column, cell.value)

    def print_body(self):

        vlist = []
        for row in self.body:
            rv = [cell.value for cell in row]
            vlist.append(rv)
        print(json.dumps(vlist, ensure_ascii=False))

    def get_title_index(self, title, first_if_has_duplicate=True):

        groups = self.group_header_cells_by_column_title()
        cells = groups.get(title, [])
        if len(cells) < 1:
            raise KeyError()
        index = 0 if first_if_has_duplicate else -1
        return cells[index].column - 1

    def print_all(self):
        n = 0
        for row in self.sheet.rows:
            n = n + 1
            print('第{}行'.format(n))
            for cell in row:
                print(cell.value)
            print('\n')

    def group_header_cells_by_column_title(self):

        group: typing.Dict[typing.Any, typing.List[Cell]] = {}
        for cell in self.head:
            name = cell.value
            if name not in group:
                group[name] = [cell]
            else:
                group[name].append(cell)
        return group

    def group_header_cells_by_column_number(self):
        """

        - Note: column number of cell (1-based)

        """

        group: typing.Dict[typing.Any, typing.List[Cell]] = {}
        for cell in self.head:
            name = cell.column
            if name not in group:
                group[name] = [cell]
            else:
                group[name].append(cell)
        return group

    def group_header_cells_by_column_letter(self):

        group: typing.Dict[typing.Any, typing.List[Cell]] = {}
        for cell in self.head:
            name = cell.column_letter
            if name not in group:
                group[name] = [cell]
            else:
                group[name].append(cell)
        return group

    def has_table_head_title(self, titles: typing.Union[typing.List[typing.Any], typing.Dict[typing.Union[int, str], typing.Any],]):

        pass


if __name__ == "__main__":
    checker = ExcelDataFinder(r"C:\Users\siwenwei\Downloads\20251107105803.xlsx")
    # crow = {"省": "广东省", "市": "东莞市"}
    # crow = {"省": re.compile("自治区$"), "乙方签约协议": "已上传"}

    def cell_matcher(k, cell):
        return True if re.compile("自治区$").search(cell.value) else False

    key_type = "column title"
    crow = {"省": cell_matcher, "乙方签约协议": "已上传"}

    key_type = "column number"
    crow = {8: cell_matcher, 40: "已上传"}

    # key_type = "column letter"
    # crow = {"H": cell_matcher, "AN": "已上传"}
    for row in checker.findall(crow, key_type=key_type):
        checker.print_table_row(row)

    # row = checker.search(crow, key_type=key_type)
    # checker.print_table_row(row)
